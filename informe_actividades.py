"""
═══════════════════════════════════════════════════════════════════════
 INFORME DE ACTIVIDADES — COLEGIO MAYOR / SED MEDELLÍN
 Documento: "Informe de actividades terminado y firmado" (GJ-FR-005)
═══════════════════════════════════════════════════════════════════════

FLUJO:
  1. Lee la cédula y número de informe → los escribe en el Excel
  2. El Excel carga todo automáticamente desde el CONSOLIDADO (VLOOKUP)
  3. Lee las evidencias del CSV → las clasifica en las actividades reales
     del contratista (leídas también del CONSOLIDADO, no hardcodeadas)
  4. Redacta metas cortas por actividad
  5. LibreOffice convierte el Excel lleno a PDF con formato oficial intacto
  6. Estampa la imagen de firma del contratista en "Firma del Contratista"
  7. Entrega: PDF firmado listo para VoBo supervisor y empaque final

APLICA PARA CUALQUIER CONTRATISTA del formato 106407:
  - Funciona independiente del número de actividades (5, 8, 10, 12...)
  - Lee actividades reales desde el CONSOLIDADO según la cédula
  - Keywords generados dinámicamente desde el texto de cada actividad

DEPENDENCIAS:
  pip install openpyxl pypdf reportlab
  apt install libreoffice

USO DESDE TERMINAL:
  python informe_actividades.py \\
    --excel  "Formato Informe MAYO - 106407 - ACTUALIZADO.xlsx" \\
    --csv    "calendario_mayo.csv" \\
    --firma  "firma.png" \\
    --cedula 1128425027 \\
    --cobro  5

USO DESDE OTRO SCRIPT:
  from informe_actividades import generar_informe
  generar_informe(
      excel_src="Formato Informe.xlsx",
      csv_evidencias="calendario.csv",
      firma_img="firma.png",
      cedula=1128425027,
      cobro=5,
      output_pdf="informe_firmado.pdf"
  )
"""

import csv
import io
import re
import shutil
import subprocess
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas

# ─────────────────────────────────────────────────────────────────────
# PALABRAS A IGNORAR EN EL CSV (spam, sistemas, notificaciones)
# ─────────────────────────────────────────────────────────────────────
SKIP_REMITENTES = [
    's@mi', 'innovación digital', 'apropia tic', 'microsoft',
    'sharepoint', 'google calendar', 'cscsoporte', 'outlook',
    'miro team', 'comunicación organizacional', 'comunicaciones educacion',
    'the miro team', 'spinach ai', 'fireflies', 'supernormal',
]
SKIP_ASUNTOS = [
    'mantenimiento', 'intermitencia', 'inactivación', 'indisponibilidad',
    'actualización de seguridad', 'migración a la nube', 'boletin',
    'encuesta de satisfacción', 'recuerde diligenciar', 'black day',
    'descuento', 'oferta', 'cancelar suscripción', 'levi',
    'resumen diario de reacciones',
]


# ═════════════════════════════════════════════════════════════════════
# PASO 0 — LEER DATOS DEL CONTRATISTA DESDE EL CONSOLIDADO
# ═════════════════════════════════════════════════════════════════════
def leer_contratista(excel_src: str, cedula: int) -> dict:
    """
    Lee nombre, contrato y actividades reales del contratista
    directamente desde la hoja CONSOLIDADO del Excel.
    Funciona para cualquier contratista independiente del número de actividades.
    """
    wb = openpyxl.load_workbook(excel_src, data_only=True, read_only=True)

    if 'CONSOLIDADO' not in wb.sheetnames:
        raise ValueError("El Excel no tiene hoja 'CONSOLIDADO'")

    wsc = wb['CONSOLIDADO']
    cedula_str = str(cedula).strip()

    for row in wsc.iter_rows(values_only=True):
        if str(row[0] or '').strip() == cedula_str:
            nombre = str(row[2] or '').strip()
            contrato = str(row[1] or '').strip()

            # Actividades en columnas 31-55 (índice 30-54), hasta encontrar vacío
            actividades = []
            for j in range(30, 55):
                if j < len(row):
                    v = row[j]
                    if v and str(v).strip():
                        actividades.append(str(v).strip())
                    else:
                        break

            wb.close()
            return {
                'nombre': nombre,
                'contrato': contrato,
                'actividades': actividades,
                'num_actividades': len(actividades),
            }

    wb.close()
    raise ValueError(f"Cédula {cedula} no encontrada en el CONSOLIDADO")


# ═════════════════════════════════════════════════════════════════════
# PASO 1 — GENERAR KEYWORDS DINÁMICAMENTE DESDE LAS ACTIVIDADES
# ═════════════════════════════════════════════════════════════════════
def extraer_keywords(actividad: str, min_len: int = 5) -> list[str]:
    """
    Extrae palabras clave relevantes del texto de una actividad.
    Ignora artículos, preposiciones y palabras muy cortas.
    """
    stop_words = {
        'de', 'la', 'el', 'en', 'los', 'las', 'del', 'al', 'con', 'por',
        'para', 'una', 'un', 'su', 'sus', 'que', 'se', 'es', 'son', 'como',
        'más', 'sus', 'las', 'los', 'nos', 'sus', 'este', 'esta', 'esto',
        'principalmente', 'orientadas', 'orientados', 'fortalecer',
    }
    texto = actividad.lower()
    palabras = re.findall(r'[a-záéíóúüñ]+', texto)
    keywords = []
    for p in palabras:
        if len(p) >= min_len and p not in stop_words:
            keywords.append(p)
    palabras_orig = actividad.lower().split()
    for i in range(len(palabras_orig) - 1):
        bigrama = palabras_orig[i] + ' ' + palabras_orig[i+1]
        if len(bigrama) >= 8:
            keywords.append(bigrama)
    return list(set(keywords))


# ═════════════════════════════════════════════════════════════════════
# PASO 2 — LEER Y CLASIFICAR EVIDENCIAS DEL CSV
# ═════════════════════════════════════════════════════════════════════
def leer_evidencias(csv_path: str) -> list[dict]:
    """Lee el CSV de calendario/correos y filtra los relevantes."""
    evidencias = []
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            asunto = row.get('Asunto', '').strip()
            remitente = row.get('De: (nombre)', '').strip().lower()
            cuerpo = row.get('Cuerpo', '').strip()[:300]

            if any(s in remitente for s in SKIP_REMITENTES):
                continue
            if any(s in asunto.lower() for s in SKIP_ASUNTOS):
                continue
            if not asunto or len(asunto) < 5:
                continue

            evidencias.append({'asunto': asunto, 'cuerpo': cuerpo})
    return evidencias


def clasificar_evidencias(evidencias: list[dict],
                          actividades: list[str]) -> list[list[str]]:
    """
    Clasifica cada evidencia en la actividad que mejor coincide.
    Funciona para cualquier número de actividades.
    """
    keywords_por_act = [extraer_keywords(act) for act in actividades]
    clasificados = [[] for _ in range(len(actividades))]

    for ev in evidencias:
        texto = (ev['asunto'] + ' ' + ev['cuerpo']).lower()
        best, best_score = 0, -1
        for i, kws in enumerate(keywords_por_act):
            score = sum(1 for k in kws if k in texto)
            if score > best_score:
                best_score = score
                best = i
        clasificados[best].append(ev['asunto'])

    return clasificados


# ═════════════════════════════════════════════════════════════════════
# PASO 3 — REDACTAR METAS CORTAS POR ACTIVIDAD
# ═════════════════════════════════════════════════════════════════════
def redactar_metas(clasificados: list[list[str]],
                   actividades: list[str]) -> list[str]:
    """
    Redacta metas profesionales usando la API de Anthropic.
    Convierte temas/evidencias en bullets formales con lenguaje institucional.
    Funciona para cualquier número de actividades (5, 8, 10, 12...).
    Fallback a redacción local si la API no responde.
    """
    import urllib.request
    import json as _json
    import os as _os

    def limpiar(asunto: str) -> str:
        return re.sub(
            r'^(RE:|RV:|FW:|Fw:|Aceptado:|Aceptada:|Cancelado:|Provisional:|Rechazado:)\s*',
            '', asunto, flags=re.IGNORECASE
        ).strip()

    def top_temas(evs: list, n: int = 8) -> list[str]:
        seen, res = set(), []
        for e in evs:
            k = limpiar(e)[:65]
            if k not in seen and len(k) > 5:
                seen.add(k)
                res.append(k)
            if len(res) >= n:
                break
        return res

    n_acts = len(actividades)

    partes = []
    for i, (act, cls) in enumerate(zip(actividades, clasificados)):
        temas = top_temas(cls, 8)
        temas_txt = ', '.join(temas) if temas else 'actividades contractuales del mes'
        partes.append(f'Actividad {i+1}: {act}\nTemas del mes: {temas_txt}')

    prompt = (
        "Eres asistente para contratistas de la Secretaría de Educación de Medellín.\n"
        "Redacta las metas del informe de ejecución contractual mensual.\n\n"
        "Para cada actividad redacta exactamente 3 bullets concretos y profesionales "
        "basados en los temas del mes.\n\n"
        "Reglas:\n"
        "- Cada bullet empieza con \"• Se \" (voz pasiva impersonal)\n"
        "- Máximo 85 caracteres por bullet\n"
        "- Lenguaje formal y específico — NO copies los temas literalmente\n"
        "- Usa verbos: asesoró, revisó, apoyó, consolidó, articuló, analizó, gestionó\n\n"
        + "\n\n".join(partes)
        + f"\n\nResponde SOLO JSON sin markdown:\n"
        + '{"metas": ["• Se ...\\n• Se ...\\n• Se ...", ...]}\n'
        + f"Array de exactamente {n_acts} elementos."
    )

    # Intentar con API de Anthropic — lista de modelos a probar en orden
    modelos = [
        'claude-sonnet-4-5-20250929',
        'claude-sonnet-4-6',
        'claude-3-5-sonnet-20241022',
    ]
    api_key = _os.environ.get('ANTHROPIC_API_KEY', '')

    if not api_key:
        print('  ⚠ ANTHROPIC_API_KEY no configurada — usando redacción local')
    else:
        for modelo in modelos:
            try:
                data = _json.dumps({
                    'model': modelo,
                    'max_tokens': 150 * n_acts,
                    'messages': [{'role': 'user', 'content': prompt}]
                }).encode('utf-8')
                req = urllib.request.Request(
                    'https://api.anthropic.com/v1/messages',
                    data=data,
                    headers={
                        'Content-Type': 'application/json',
                        'x-api-key': api_key,
                        'anthropic-version': '2023-06-01'
                    },
                    method='POST'
                )
                with urllib.request.urlopen(req, timeout=45) as resp:
                    result = _json.loads(resp.read())
                txt = (result['content'][0]['text']
                       .strip().replace('```json', '').replace('```', '').strip())
                metas = _json.loads(txt)['metas']
                if len(metas) == n_acts:
                    print(f'  ✓ Metas redactadas por IA ({modelo}) para {n_acts} actividades')
                    return metas
            except Exception as e:
                print(f'  ⚠ Modelo {modelo} falló: {e}')
                continue
        print('  ⚠ Ningún modelo de la API respondió — usando redacción local')

    # Fallback: redacción local sin API
    metas = []
    for evs, act in zip(clasificados, actividades):
        temas = top_temas(evs, 3)
        if temas:
            bullets = '\n'.join(f'• Se gestionaron acciones relacionadas con: {t}.' for t in temas)
        else:
            verbo = act.split()[0].lower() if act else 'cumplir'
            bullets = f'• Se apoyaron las actividades establecidas en el contrato.'
        metas.append(bullets)
    return metas


# ═════════════════════════════════════════════════════════════════════
# PASO 4 — LLENAR EL EXCEL
# ═════════════════════════════════════════════════════════════════════
def llenar_excel(excel_src: str, cedula: int, cobro: int,
                 metas: list[str], excel_dst: str) -> None:
    """
    Copia el Excel original y escribe SOLO:
      - G14 = cédula  → activa todos los VLOOKUP del CONSOLIDADO
      - E12 = número de informe
      - E16 ... E(15+n) = meta por actividad (una fila por actividad)
    El Excel calcula todo lo demás automáticamente.
    """
    shutil.copy(excel_src, excel_dst)
    wb = openpyxl.load_workbook(excel_dst)
    ws = wb['INFORME']

    ws['G14'] = cedula
    ws['E12'] = cobro

    COL_CHARS = 88
    FONT_SIZE = 10

    for i, meta in enumerate(metas):
        fila = 16 + i
        cell = ws[f'E{fila}']
        cell.value = meta
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        cell.font = Font(name='Calibri', size=FONT_SIZE)

        lineas_total = sum(max(1, -(-len(b) // COL_CHARS)) for b in meta.splitlines())
        altura = max(144, lineas_total * 15 + 10)
        ws.row_dimensions[fila].height = altura

    wb.save(excel_dst)
    print(f'  ✓ Excel llenado: G14={cedula}, E12={cobro}, '
          f'metas en E16:E{15+len(metas)}')


# ═════════════════════════════════════════════════════════════════════
# PASO 5 — CONVERTIR EXCEL A PDF CON LIBREOFFICE
# ═════════════════════════════════════════════════════════════════════
def excel_a_pdf(excel_path: str, output_dir: str) -> str:
    """
    Convierte el Excel a PDF usando LibreOffice.
    Respeta el formato oficial intacto: logo, tablas, fórmulas calculadas.
    """
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'pdf',
         excel_path, '--outdir', output_dir],
        capture_output=True, text=True, timeout=100
    )
    if result.returncode != 0:
        raise RuntimeError(f'LibreOffice falló: {result.stderr}')

    pdf_path = str(Path(output_dir) / (Path(excel_path).stem + '.pdf'))
    print(f'  ✓ Convertido a PDF: {pdf_path}')
    return pdf_path


# ═════════════════════════════════════════════════════════════════════
# PASO 6 — ESTAMPAR FIRMA EN "FIRMA DEL CONTRATISTA" (última página)
# ═════════════════════════════════════════════════════════════════════
def estampar_firma(pdf_src: str, firma_img: str,
                   pdf_dst: str, cedula: int, nombre: str) -> None:
    """
    Estampa la imagen de firma (PNG/JPG dibujada o escaneada) en el
    campo 'Firma del Contratista' de la última página del informe.

    Coordenadas medidas del PDF real (595x841pt):
      Línea de firma en y=515.6 desde arriba
      Espacio libre: y=479.9 → y=515.6 (35.7pt)
    """
    reader = PdfReader(pdf_src)
    writer = PdfWriter()
    total = len(reader.pages)

    for i in range(total - 1):
        writer.add_page(reader.pages[i])

    last = reader.pages[-1]
    page_w = float(last.mediabox.width)
    page_h = float(last.mediabox.height)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(page_w, page_h))

    firma_path = Path(firma_img) if firma_img else None

    if firma_path and firma_path.exists():
        firma_w, firma_h = 165, 32
        firma_x = 42
        firma_y = page_h - 515.6
        c.drawImage(
            str(firma_path), firma_x, firma_y,
            width=firma_w, height=firma_h,
            preserveAspectRatio=True, mask='auto'
        )
        print(f'  ✓ Firma imagen estampada desde: {firma_path}')
    else:
        c.setFont('Helvetica-Oblique', 11)
        c.setFillColorRGB(0.05, 0.05, 0.4)
        c.drawString(60, page_h - 505, nombre)
        c.setFont('Helvetica', 7)
        c.setFillColorRGB(0.3, 0.3, 0.3)
        c.drawString(60, page_h - 516, f'C.C. {cedula}')
        print(f'  ⚠ Sin imagen de firma — usando texto como fallback')

    c.save()
    buf.seek(0)

    overlay = PdfReader(buf).pages[0]
    last.merge_page(overlay)
    writer.add_page(last)

    with open(pdf_dst, 'wb') as f:
        writer.write(f)
    print(f'  ✓ PDF firmado guardado: {pdf_dst}')


# ═════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL — ORQUESTA TODO EL PROCESO
# ═════════════════════════════════════════════════════════════════════
def generar_informe(
    excel_src: str,
    csv_evidencias: str,
    cedula: int,
    cobro: int,
    firma_img: str = '',
    output_pdf: str = '',
    workdir: str = '/tmp',
) -> str:
    """
    Genera el 'Informe de actividades terminado y firmado'.
    Aplica para cualquier contratista del formato 106407.
    """
    mes_label = f'Informe{cobro}'
    if not output_pdf:
        output_pdf = str(Path(workdir) / f'{cedula}_{mes_label}_FIRMADO.pdf')

    print(f'\n{"═"*60}')
    print(f'  INFORME DE ACTIVIDADES — Cédula {cedula} · Cobro #{cobro}')
    print(f'{"═"*60}')

    print('\n[1/5] Leyendo datos del contratista desde CONSOLIDADO...')
    datos = leer_contratista(excel_src, cedula)
    print(f'  Contratista: {datos["nombre"]}')
    print(f'  Contrato:    {datos["contrato"]}')
    print(f'  Actividades: {datos["num_actividades"]}')
    for i, a in enumerate(datos['actividades']):
        print(f'    Act {i+1}: {a[:70]}...' if len(a) > 70 else f'    Act {i+1}: {a}')

    print(f'\n[2/5] Leyendo evidencias desde {csv_evidencias}...')
    evidencias = leer_evidencias(csv_evidencias) if csv_evidencias else []
    print(f'  {len(evidencias)} evidencias relevantes encontradas')
    clasificados = clasificar_evidencias(evidencias, datos['actividades'])
    for i, evs in enumerate(clasificados):
        print(f'  Act {i+1}: {len(evs)} evidencias')

    print('\n[3/5] Redactando metas por actividad...')
    metas = redactar_metas(clasificados, datos['actividades'])

    print('\n[4/5] Llenando Excel...')
    excel_tmp = str(Path(workdir) / f'{cedula}_Informe{cobro}_tmp.xlsx')
    llenar_excel(excel_src, cedula, cobro, metas, excel_tmp)

    pdf_tmp = excel_a_pdf(excel_tmp, workdir)

    print('\n[5/5] Estampando firma...')
    estampar_firma(pdf_tmp, firma_img, output_pdf, cedula, datos['nombre'])

    print(f'\n{"═"*60}')
    print(f'  ✅ LISTO: {output_pdf}')
    print(f'  → Informe de actividades terminado y firmado')
    print(f'  → Listo para VoBo del supervisor y empaque final')
    print(f'{"═"*60}\n')

    return output_pdf


# ═════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Genera el informe de actividades firmado (GJ-FR-005). '
                    'Aplica para cualquier contratista del formato 106407.'
    )
    parser.add_argument('--excel',   required=True,
                        help='Ruta al Excel del formato (con hoja CONSOLIDADO)')
    parser.add_argument('--csv',     required=True,
                        help='Ruta al CSV de evidencias (calendario/correos del mes)')
    parser.add_argument('--cedula',  required=True, type=int,
                        help='Número de cédula del contratista')
    parser.add_argument('--cobro',   required=True, type=int,
                        help='Número de informe/cobro')
    parser.add_argument('--firma',   default='',
                        help='Ruta a imagen PNG/JPG de la firma (dibujada o escaneada)')
    parser.add_argument('--output',  default='',
                        help='Ruta del PDF de salida')
    parser.add_argument('--workdir', default='/tmp',
                        help='Directorio temporal (default: /tmp)')
    args = parser.parse_args()

    generar_informe(
        excel_src=args.excel,
        csv_evidencias=args.csv,
        cedula=args.cedula,
        cobro=args.cobro,
        firma_img=args.firma,
        output_pdf=args.output,
        workdir=args.workdir,
    )
